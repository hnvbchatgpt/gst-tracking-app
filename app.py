from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import json
import os
from datetime import datetime
from config import Config
from models import Client, GSTReturn
from database import create_database_tables
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename
import tempfile

app = Flask(__name__)
app.config.from_object(Config)

# Ensure database tables exist
create_database_tables()

# Initialize models
client_model = Client()
gst_return_model = GSTReturn()

@app.route('/')
def index():
    """Main dashboard page"""
    return render_template('index.html')

@app.route('/master_data')
def master_data():
    """Master data management page"""
    clients = client_model.get_all_clients()
    return render_template('master_data.html', 
                         clients=clients,
                         taxpayer_types=Config.TAXPAYER_TYPES)

@app.route('/gst_returns')
def gst_returns():
    """GST returns management page"""
    return render_template(
        'gst_returns.html',
        returns=Config.GST_RETURNS,
        financial_years=Config.get_financial_years(),
        months=Config.MONTHS,
        quarters=Config.QUARTERS,
        gst_returns_json=json.dumps(Config.GST_RETURNS)
    )

@app.route('/api/clients', methods=['GET'])
def get_clients():
    """API endpoint to get all clients"""
    try:
        clients = client_model.get_all_clients()
        clients_data = []
        for client in clients:
            clients_data.append({
                'client_code': client[0],
                'client_name': client[1],
                'date_of_registration': client[2].strftime('%Y-%m-%d') if client[2] else None,
                'effective_date_of_cancellation': client[3].strftime('%Y-%m-%d') if client[3] else None,
                'gstin': client[4],
                'taxpayer_type': client[5],
                'gst_portal_userid': client[6],
                'gst_portal_password': client[7],
                'eway_bill_userid': client[8],
                'eway_bill_password': client[9],
                'client_email_id': client[10],
                'mobile_no': client[11],
                'email_password': client[12]
            })
        return jsonify({'success': True, 'data': clients_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clients', methods=['POST'])
def create_client():
    """API endpoint to create new client"""
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['client_name', 'date_of_registration', 'gstin', 
                          'taxpayer_type', 'gst_portal_userid', 'gst_portal_password',
                          'client_email_id', 'mobile_no']
        
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'})
        
        # Convert date strings to date objects
        if data.get('date_of_registration'):
            data['date_of_registration'] = datetime.strptime(data['date_of_registration'], '%Y-%m-%d').date()
        if data.get('effective_date_of_cancellation'):
            data['effective_date_of_cancellation'] = datetime.strptime(data['effective_date_of_cancellation'], '%Y-%m-%d').date()
        
        success = client_model.create_client(data)
        
        if success:
            return jsonify({'success': True, 'message': 'Client created successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to create client'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clients/<int:client_code>', methods=['PUT'])
def update_client(client_code):
    """API endpoint to update client"""
    try:
        data = request.json
        
        # Convert date strings to date objects
        if data.get('date_of_registration'):
            data['date_of_registration'] = datetime.strptime(data['date_of_registration'], '%Y-%m-%d').date()
        if data.get('effective_date_of_cancellation'):
            data['effective_date_of_cancellation'] = datetime.strptime(data['effective_date_of_cancellation'], '%Y-%m-%d').date()
        
        success = client_model.update_client(client_code, data)
        
        if success:
            return jsonify({'success': True, 'message': 'Client updated successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update client'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clients/<int:client_code>', methods=['DELETE'])
def delete_client(client_code):
    """API endpoint to delete client"""
    try:
        success = client_model.delete_client(client_code)
        
        if success:
            return jsonify({'success': True, 'message': 'Client deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to delete client'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/return_dashboard', methods=['POST'])
def get_return_dashboard():
    try:
        data = request.json
        frequency = data.get('frequency')
        financial_year = data.get('financial_year')
        month = data.get('month')
        quarter = data.get('quarter')

        MONTH_SHORT_MAP = {
            "January": "Jan", "February": "Feb", "March": "Mar",
            "April": "Apr", "May": "May", "June": "Jun",
            "July": "Jul", "August": "Aug", "September": "Sep",
            "October": "Oct", "November": "Nov", "December": "Dec",
            # Also include 3-letter codes for idempotency
            "Jan": "Jan", "Feb": "Feb", "Mar": "Mar", "Apr": "Apr", "May": "May",
            "Jun": "Jun", "Jul": "Jul", "Aug": "Aug", "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dec": "Dec"
        }

        def get_period_year_from_financial_year(month, financial_year):
            fy_start_year = int(financial_year.split('-')[0])
            if month in ['Jan', 'Feb', 'Mar']:
                return fy_start_year + 1  # Jan-Mar: next calendar year
            else:
                return fy_start_year
        
        def get_period_year_from_financial_year_for_quartely(quarter, financial_year):
            fy_start_year = int(financial_year.split('-')[0])
            if quarter in ['Jan-Mar']:
                return fy_start_year + 1  # Jan-Mar: next calendar year
            else:
                return fy_start_year

        # Convert month to three-letter code
        month_short = MONTH_SHORT_MAP.get(month, month)

        if frequency == 'Monthly':
            period_year = get_period_year_from_financial_year(month_short, financial_year)
            period = f"{month_short}-{period_year}"
        elif frequency == 'Quarterly':
            month_short = MONTH_SHORT_MAP.get(quarter[-3:], quarter[-3:])
            period_year = get_period_year_from_financial_year(month_short, financial_year)
            period = f"{month_short}-{period_year}"
        elif frequency == 'Annually':
            period = financial_year
        else:
            return jsonify({'success': False, 'error': 'Invalid frequency'})

        dashboard_data = {}
        for return_type, return_config in Config.GST_RETURNS.items():
            if return_config['frequency'] == frequency:
                dashboard_data[return_type] = gst_return_model.get_return_dashboard_data(return_type, period)

        return jsonify({'success': True, 'data': dashboard_data, 'period': period})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
        

@app.route('/api/return_clients', methods=['POST'])
def get_return_clients():
    """API endpoint to get clients for specific return type and period"""
    try:
        data = request.json
        return_type = data.get('return_type')
        period = data.get('period')
        
        # Get applicable clients
        clients = gst_return_model.get_applicable_clients(return_type, period)
        
        clients_data = []
        for client in clients:
            # Get return data if exists
            return_data = gst_return_model.get_return_data(client[0], return_type, period)
            
            client_info = {
                'client_code': client[0],
                'client_name': client[1],
                'gstin': client[2],
                'period': period,
                'date_of_filing': return_data[4].strftime('%Y-%m-%d') if return_data and return_data[4] else None,
                'status': return_data[5] if return_data else 'Data Received',
                'arn': return_data[6] if return_data else None,
                'remarks': return_data[7] if return_data else None
            }
            clients_data.append(client_info)
        
        return jsonify({'success': True, 'data': clients_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/save_return_data', methods=['POST'])
def save_return_data():
    """API endpoint to save return data"""
    try:
        data = request.json
        
        # Convert date string to date object
        if data.get('date_of_filing'):
            data['date_of_filing'] = datetime.strptime(data['date_of_filing'], '%Y-%m-%d').date()
        
        # ✅ Validation: ARN required if status is Filed
        if data.get('status') == 'Filed':
            missing_fields = []
            if not data.get('arn'):
                missing_fields.append('ARN')
            if not data.get('date_of_filing'):
                missing_fields.append('Date of Filing')

            if missing_fields:
                missing_str = ' and '.join(missing_fields)
                return jsonify({
                    'success': False, 
                    'error': f'{missing_str} required.'
                })
        
        success = gst_return_model.save_return_data(data)
        
        if success:
            return jsonify({'success': True, 'message': 'Return data saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save return data'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/export_clients')
def export_clients():
    """Export clients to Excel"""
    try:
        clients = client_model.get_all_clients()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Client Master Data"
        
        # Define headers
        headers = [
            'Client Code', 'Client Name*', 'Date of Registration*', 'Effective Date of Cancellation',
            'GSTIN*', 'Taxpayer Type*', 'GST Portal User ID*', 'GST Portal Password*',
            'EWAY Bill User ID', 'EWAY Bill Password', 'Client Email ID*', 'Mobile No*', 'Email Password'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        mandatory_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            if '*' in header:
                cell.fill = mandatory_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add client data
        for row_num, client in enumerate(clients, 2):
            ws.cell(row=row_num, column=1, value=client[0])
            ws.cell(row=row_num, column=2, value=client[1])
            ws.cell(row=row_num, column=3, value=client[2].strftime('%Y-%m-%d') if client[2] else None)
            ws.cell(row=row_num, column=4, value=client[3].strftime('%Y-%m-%d') if client[3] else None)
            ws.cell(row=row_num, column=5, value=client[4])
            ws.cell(row=row_num, column=6, value=client[5])
            ws.cell(row=row_num, column=7, value=client[6])
            ws.cell(row=row_num, column=8, value=client[7])
            ws.cell(row=row_num, column=9, value=client[8])
            ws.cell(row=row_num, column=10, value=client[9])
            ws.cell(row=row_num, column=11, value=client[10])
            ws.cell(row=row_num, column=12, value=client[11])
            ws.cell(row=row_num, column=13, value=client[12])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(temp_file.name, 
                        as_attachment=True, 
                        download_name=f'client_master_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/import_clients', methods=['POST'])

def import_clients():
    """Import clients from Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file selected for upload.'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Filename is empty.'})

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload a .xlsx or .xls Excel file.'})

        # Save uploaded Excel file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()  # Crucial to prevent [WinError 32]

        # Load workbook
        wb = openpyxl.load_workbook(temp_file.name, data_only=True)
        ws = wb.active

        imported_count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]  # Read header row for reference

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):
                    continue  # Skip completely empty rows

                client_data = {
                    'client_name': row[1],
                    'date_of_registration': row[2],
                    'effective_date_of_cancellation': row[3],
                    'gstin': row[4],
                    'taxpayer_type': str(row[5]).strip() if row[5] else None,
                    'gst_portal_userid': row[6],
                    'gst_portal_password': row[7],
                    'eway_bill_userid': row[8],
                    'eway_bill_password': row[9],
                    'client_email_id': row[10],
                    'mobile_no': str(row[11]).strip() if row[11] else None,
                    'email_password': row[12]
                }

                # Check required fields
                required_fields = [
                    'client_name', 'date_of_registration', 'gstin',
                    'taxpayer_type', 'gst_portal_userid', 'gst_portal_password',
                    'client_email_id', 'mobile_no'
                ]
                missing_fields = [field for field in required_fields if not client_data.get(field)]

                if missing_fields:
                    errors.append(f"Row {row_num}: Missing required fields: {', '.join(missing_fields)}")
                    continue

                # Convert dates (Excel can return date objects or strings)
                if isinstance(client_data['date_of_registration'], str):
                    try:
                        client_data['date_of_registration'] = datetime.strptime(
                            client_data['date_of_registration'], '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid date format for 'Date of Registration'")
                        continue

                if client_data['effective_date_of_cancellation']:
                    if isinstance(client_data['effective_date_of_cancellation'], str):
                        try:
                            client_data['effective_date_of_cancellation'] = datetime.strptime(
                                client_data['effective_date_of_cancellation'], '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid date format for 'Effective Date of Cancellation'")
                            continue

                # Create client
                created = client_model.create_client(client_data)
                if created:
                    imported_count += 1
                else:
                    errors.append(f"Row {row_num}: Failed to create client (possibly duplicate GSTIN or DB error)")

            except Exception as ex:
                errors.append(f"Row {row_num}: Unexpected error: {str(ex)}")

        # Clean up temp file
        os.unlink(temp_file.name)

        return jsonify({
            'success': True,
            'imported_count': imported_count,
            'errors': errors
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f"Import failed: {str(e)}"})


@app.route('/api/download_template')
def download_template():
    """Download Excel template for client import"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Client Master Template"
        
        # Define headers
        headers = [
            'Client Code (Auto Generated)', 'Client Name*', 'Date of Registration* (YYYY-MM-DD)', 
            'Effective Date of Cancellation (YYYY-MM-DD)', 'GSTIN*', 'Taxpayer Type*', 
            'GST Portal User ID*', 'GST Portal Password*', 'EWAY Bill User ID', 
            'EWAY Bill Password', 'Client Email ID*', 'Mobile No*', 'Email Password'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        mandatory_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            if '*' in header:
                cell.fill = mandatory_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add sample data
        sample_data = [
            ['', 'ABC Industries Ltd', '2024-04-01', '', '27AAAAA0000A1Z5', 'Monthly', 
             'abc@gst.gov.in', 'password123', 'abc@ewaybill.nic.in', 'ewaypass123', 
             'abc@company.com', '9876543210', 'emailpass123']
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Add instructions
        instructions = [
            "INSTRUCTIONS:",
            "1. Fields marked with * are mandatory",
            "2. Date format should be YYYY-MM-DD (e.g., 2024-04-01)",
            "3. Taxpayer Type should be one of: Monthly, Quarterly, Composition",
            "4. GSTIN should be 15 characters",
            "5. Do not modify the Client Code column - it will be auto-generated",
            "6. Remove this instruction section before importing"
        ]
        
        start_row = len(sample_data) + 4
        for i, instruction in enumerate(instructions):
            ws.cell(row=start_row + i, column=1, value=instruction)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(temp_file.name, 
                        as_attachment=True, 
                        download_name='client_master_template.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
